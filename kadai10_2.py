import matplotlib
import matplotlib.pyplot as plt
import japanize_matplotlib
import numpy as np
import pandas as pd
import datetime
import sklearn
import openpyxl
from sklearn.tree import DecisionTreeClassifier, plot_tree
from sklearn.ensemble  import RandomForestClassifier
from sklearn.metrics import f1_score

weather_df = pd.read_excel('ds07_temp_power_exercise.xlsx', 
                            sheet_name='2020熊谷市気象データ_data', skiprows=[0,1,2],
                            names=['年月日', '平均気温', '', '', '最高気温', '', '', '', '', '最低気温', '', '', '', '', 
                            '降水量', '', '', '', '日照時間', '', '', '', '降雪量', '', '', '', '平均風速', '', '', 
                            '平均蒸気圧', '', '', '平均湿度', '', '', '平均現地気圧', '', '', '', '', '', '天気概況', '', ''])

year = 2020
month = 1
data = ['平均気温', '最高気温', '最低気温', '降水量', '日照時間', '降雪量', '平均風速', '平均蒸気圧', '平均湿度', '平均現地気圧'] 
target = '天気概況'
weather_data = weather_df[(datetime.datetime(year,month,1) <= weather_df['年月日']) 
                        & (weather_df['年月日'] < datetime.datetime(year,month+1,1))][data].values
weather_target = weather_df[(datetime.datetime(year,month,1) <= weather_df['年月日']) 
                            & (weather_df['年月日'] < datetime.datetime(year,month+1,1))][target].values

book=openpyxl.load_workbook('kadai10_table.xlsx')
sheet1=book['Sheet1']
sheet2=book['Sheet2']

def decision_tree_f1(end_depth):
    for max_d in range(1,end_depth+1):
        sheet1.cell(row=max_d+1, column=2).value=max_d
        clf = DecisionTreeClassifier(criterion='entropy', max_depth=max_d)
        clf = clf.fit(weather_data, weather_target)
        _f1_score=f1_score(weather_target, clf.predict(weather_data), average="micro")
        sheet1.cell(row=max_d+1, column=3).value=_f1_score
        # print('max_depth='+str(max_d))
        # print('f1_score:', _f1_score)

def random_forest_f1(end_depth, end_estimator):
    # ランダムフォレストを作成してF値を計算 (木の数、木の深さはグリッドサーチ)
    best_f1_score = 0
    best_estimator = 0
    best_depth = 0
    for _depth in range(1,end_depth+1):
        # print('max_depth='+str(_depth))
        sheet2.cell(row=_depth+2, column=2).value=_depth
        for _estimator in range(1,end_estimator+1):
            rf = RandomForestClassifier(n_estimators=_estimator, max_depth=_depth, random_state=1)
            rf = rf.fit(weather_data, weather_target)
            _f1_score = f1_score(weather_target, rf.predict(weather_data), average="micro")
            
            sheet2.cell(row=2, column=_estimator+2).value=_estimator
            sheet2.cell(row=_depth+2, column=_estimator+2).value=_f1_score
            if _f1_score > best_f1_score:
                best_f1_score = _f1_score
                best_estimator = _estimator
                best_depth = _depth
    print(f'best n_estimators={best_estimator}')
    print(f'best depth={best_depth}')
    print(f'best f1_score={best_f1_score}')
        #     print('n_estimators='+str(_estimator))
        #     print('f1_score: '+str(_f1_score))
        # print()

def feature_importances_calculator(n_est,max_d):
    # 説明変数の重要度をグラフに出力
    rf = RandomForestClassifier(n_estimators=n_est, max_depth=max_d, random_state=1)
    rf = rf.fit(weather_data, weather_target)
    importances = rf.feature_importances_
    plt.figure(figsize=(8, 3))
    plt.barh(range(len(data)), rf.feature_importances_ , align='center')
    plt.yticks(np.arange(len(data)), data)
    plt.show()

#print('決定木の深さと精度')
#decision_tree_f1(end_depth=8)

#print()

#print('ランダムフォレストの木の数・深さと精度')
#random_forest_f1(end_depth=8, end_estimator=20)

feature_importances_calculator(n_est=18,max_d=5)

# book.save('kadai10_table_rev.xlsx')
# book.close()